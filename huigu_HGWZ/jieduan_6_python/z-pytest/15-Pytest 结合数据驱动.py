# encoding='utf8'
# file->settings->Editor->file and Code Templates 
# -> Python Script 输入自己的模板
# 查询输入 type check 可对应修改错误等级 例：error warning
# 当前时间 2022/12/8 12:52
''
'''Pytest 结合数据驱动-file_yaml
    大纲
        数据驱动简介
        数据驱动案例
    
    简介
        数据驱动就是数据的改变从而驱动自动化测试的执行，最终引起测试结果的改变。简单来说，
        就是参数化的应用。数据量小的测试用例可以使用代码的参数化来实现数据驱动，数据量大的情况
        下建议大家使用一种结构化的文件（例如：file_yaml、json）来对数据进行存储，然后在测试用例中读取这些数据；
        
    应用场景
        app web 接口自动化测试
        测试步骤的数据驱动
        测试数据的数据驱动
        配置的数据驱动
    
    class Test_yaml:
        @pytest.mark.parametrize('env',yaml.safe_load(open('./env.yaml',encoding='utf-8')))
        def test_yaml(self,env):
            if 'test' in env:
                print('测试环境')
                print('ip是：{}'.format(env['test']))
            elif 'develop' in env:
                print('开发环境')
                print('ip是：{}'.format(env['develop']))
                print(env)
            else:
                print('正式环境')
    
        def test_yaml_file(self):
            print(yaml.safe_load(open('./env.yaml',encoding='utf-8')))
    '''

'''Pytest 结合数据驱动-excel
    读取 Excel 文件
        第三方库
        xlrd
        xlwings
        pandas
        openpyxl
        官方文档： https://openpyxl.readthedocs.io/en/stable/
        
    openpyxl 库的安装
        安装：pip install openpyxl
        导入：import openpyxl
    
    openpyxl 库的操作
            读取工作簿
            读取工作表
            读取单元格
            import openpyxl
            # 获取工作簿
            book = openpyxl.load_workbook('../data/params.xlsx')
            # 读取工作表
            sheet = book.active
            # 读取单个单元格
            cell_a1 = sheet['A1']
            cell_a3 = sheet.cell(column=1, row=3)  # A3
            # 读取多个连续单元格
            cells = sheet["A1":"C3"]
            # 获取单元格的值
            cell_a1.value
    
    工程目录结构
        data 目录：存放 excel 数据文件
        func 目录：存放被测函数文件
        testcase 目录：存放测试用例文件
        # 工程目录结构
        .
        ├── data
        │   └── params.excel
        ├── func
        │   ├── __init__.py
        │   └── operation.py
        └── testcase
            ├── __init__.py
            └── test_add.py
    
    测试准备
        被测对象：operation.py
        测试用例：test_add.py
        # operation.py 文件内容
        def my_add(x, y):
            result = x + y
            return result
        
        # test_add.py 文件内容
        class TestWithEXCEL:
            @pytest.mark.parametrize('x,y,expected', get_excel())
            def test_add(self, x, y, expected):
                assert my_add(int(x), int(y)) == int(expected)
    测试准备
        测试数据：params.xlsx
    
    Pytest 数据驱动结合 Excel 文件
        # 读取Excel文件
        import openpyxl
        import pytest
        
        def get_excel():
            # 获取工作簿
            book = openpyxl.load_workbook('../data/params.xlsx')
        
            # 获取活动行（非空白的）
            sheet = book.active
        
            # 提取数据，格式：[[1, 2, 3], [3, 6, 9], [100, 200, 300]]
            values = []
            for row in sheet:
                line = []
                for cell in row:
                    line.append(cell.value)
                values.append(line)
            return values
'''

'''Pytest 结合数据驱动-csv
    csv 文件介绍
        csv：逗号分隔值
        是 Comma-Separated Values 的缩写
        以纯文本形式存储数字和文本
        文件由任意数目的记录组成
        每行记录由多个字段组成
        
        Linux从入门到高级,linux,¥5000
        web自动化测试进阶,python,¥3000
        app自动化测试进阶,python,¥6000
        Docker容器化技术,linux,¥5000
        测试平台开发与实战,python,¥8000
    
    csv 文件使用
        读取数据
        内置函数：open()
        内置模块：csv
        方法：csv.reader(iterable)
        
        参数：iterable ,文件或列表对象
        返回：迭代器，每次迭代会返回一行数据。
        # 读取csv文件内容
        
        def get_csv():
            with open('demo.csv', 'r') as file:
                raw = csv.reader(file)
        
                for line in raw:
                    print(line)
    
    工程目录结构
        data 目录：存放 csv 数据文件
        func 目录：存放被测函数文件
        testcase 目录：存放测试用例文件
        # 工程目录结构
        .
        ├── data
        │   └── params.csv
        ├── func
        │   ├── __init__.py
        │   └── operation.py
        └── testcase
            ├── __init__.py
            └── test_add.py
    
    测试准备
        被测对象：operation.py
        测试用例：test_add.py
        测试数据：params.csv
        # operation.py 文件内容
        def my_add(x, y):
            result = x + y
            return result
        
        # test_add.py 文件内容
        class TestWithCSV:
            @pytest.mark.parametrize('x,y,expected', [[1, 1, 2]])
            def test_add(self, x, y, expected):
                assert my_add(int(x), int(y)) == int(expected)
        
        # params.csv 文件内容
        1,1,2
        3,6,9
        100,200,300
    
    Pytest 数据驱动结合 csv 文件
            # 读取 data目录下的 params.csv 文件
            import csv
            
            def get_csv():
                """
                获取csv数据
                :return: 返回数据的结构：[[1, 1, 2], [3, 6, 9], [100, 200, 300]]
                """
                with open('../data/params.csv', 'r') as file:
                    raw = csv.reader(file)
                    data = []
                    for line in raw:
                        data.append(line)
                return data
    '''

'''Pytest 结合数据驱动-json
    json 文件介绍
        json 是 JS 对象
        全称是 JavaScript Object Notation
        是一种轻量级的数据交换格式
        json 结构
        对象 {"key": value}
        数组 [value1, value2 ...]
        {
          "name:": "hogwarts ",
          "detail": {
            "course": "python",
            "city": "北京"
          },
          "remark": [1000, 666, 888]
        }
    
    json 文件使用
        查看 json 文件
            pycharm
            txt 记事本
        读取 json 文件
            内置函数 open()
            内置库 json
        方法：json.loads()
        方法：json.dumps()
        # 读取json文件内容
        def get_json():
            with open('demo.json', 'r') as f:
                data = json.loads(f.read())
                print(data)
    
    工程目录结构
        data 目录：存放 json 数据文件
        func 目录：存放被测函数文件
        testcase 目录：存放测试用例文件
        # 工程目录结构
        .
        ├── data
        │   └── params.json
        ├── func
        │   ├── __init__.py
        │   └── operation.py
        └── testcase
            ├── __init__.py
            └── test_add.py
    
    测试准备
        被测对象：operation.py
        测试用例：test_add.py
        测试数据：params.json
        # operation.py 文件内容
        def my_add(x, y):
            result = x + y
            return result
        
        # test_add.py 文件内容
        class TestWithJSON:
            @pytest.mark.parametrize('x,y,expected', [[1, 1, 2]])
            def test_add(self, x, y, expected):
                assert my_add(int(x), int(y)) == int(expected)
        
        # params.json 文件内容
        {
          "case1": [1, 1, 2],
          "case2": [3, 6, 9],
          "case3": [100, 200, 300]
        }
    
    Pytest 数据驱动结合 json 文件
        # 读取json文件
        def get_json():
            """
            获取json数据
            :return: 返回数据的结构：[[1, 1, 2], [3, 6, 9], [100, 200, 300]]
            """
            with open('../data/params.json', 'r') as f:
                data = json.loads(f.read())
                return list(data.values())
    
    '''
#todo lianxi 目录下 data_drive中有示例
